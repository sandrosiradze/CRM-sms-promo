import os, re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook

# ---------- CONFIG ----------
BASE_DIR       = os.path.join(os.getcwd(), "daily")
LOG_DIR        = os.path.join(BASE_DIR, "logs")
PROMODATE_DIR  = BASE_DIR
PROMO_FILE     = os.path.join(os.getcwd(), "5.xlsx")   # adjust if needed
KEY            = 128
# ----------------------------

# ======= HELPERS =======
def now_strings():
    now = datetime.now()
    return now.strftime("%m.%d.%y"), now.strftime("%m.%d-%H%M%S")

def ensure_dirs(run_dir: str):
    os.makedirs(LOG_DIR, exist_ok=True)
    os.makedirs(run_dir, exist_ok=True)

def write_log(msg: str):
    day = datetime.now().strftime("%m.%d")
    logfile = os.path.join(LOG_DIR, f"log-{day}.txt")
    with open(logfile, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now().isoformat(timespec='seconds')}: {msg}\n")

def canonize(name: str) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(name).lower())

def resolve_column(actual_cols, candidates):
    cmap = {canonize(c): c for c in actual_cols}
    for cand in candidates:
        key = canonize(cand)
        if key in cmap:
            return cmap[key]
    for cand in candidates:
        for k, v in cmap.items():
            if cand in k:
                return v
    raise KeyError(f"Missing column: {candidates}")

def sanitize_sheet(name, fallback="sheet"):
    s = str(name) if pd.notna(name) and str(name).strip() else fallback
    return re.sub(r'[:\\/?*\[\]]', '-', s)[:31] or fallback

def pct_number(val):
    if pd.isna(val): return None
    s = str(val).strip()
    try:
        if s.endswith('%'): s = s[:-1].strip()
        num = float(s)
        if 0 <= num <= 1:  # allow 0.2 → 20%
            num *= 100
        return num
    except:
        return None

def pct_label(num):
    if num is None: return ""
    return f"{int(round(num))}%"

def normalize_percent(val) -> str:
    num = pct_number(val)
    if num is None:
        s = (str(val) if val is not None else "").strip()
        return s if s.endswith("%") and len(s) >= 2 else ""
    return pct_label(num)

def normalize_amount_to_lari(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    if s.endswith("₾") and len(s) >= 2:
        return s
    m = re.search(r'(\d+)', s)
    if not m:
        return ""
    return f"{m.group(1)}₾"

# ====== ENCODER (React port) ======
def simple_encrypt(text: str, key: int = 128) -> str:
    out = []
    for i, ch in enumerate(text):
        out.append(f"{ord(ch)+key+i:04x}")
    return "-".join(out)

def build_offer_url(locale: str, amount: str, percent: str, key: int = 128) -> str:
    loc = (locale or "ka").lower()
    if loc not in {"ka","tr","ru","en"}: loc = "ka"
    base = f"https://www.ambassadoribet.com/{loc}/personal-offer"
    return f"{base}?offam={simple_encrypt(amount, key)}&offperc={simple_encrypt(percent, key)}"

# ===== EXPORTS =====
def export_percentages(df, pct_col, run_dir):
    _, tag = now_strings()
    out_path = os.path.join(run_dir, f"Percentages-{tag}.xlsx")

    df = df.copy()
    df["_pct_num"] = df[pct_col].apply(pct_number)
    groups = {pct_label(n) if pct_label(n) else "Unknown%": df[df["_pct_num"]==n] 
              for n in sorted({n for n in df["_pct_num"] if n is not None})}
    if not groups:
        groups = {"Unknown%": df[df["_pct_num"].isna()]}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for label, sub in groups.items():
            sheet = sanitize_sheet(label)
            temp = sub[["userid"]].copy()
            temp["userid"] = temp["userid"].apply(lambda x: f"{x}," if pd.notna(x) else "")
            temp.to_excel(writer, sheet_name=sheet, index=False)

    write_log(f"Percentages workbook created: {out_path}")

def export_sms(df, nickname, phone, reqdep, coin, locale, pct, run_dir):
    _, tag = now_strings()
    out_path = os.path.join(run_dir, f"SMS-{tag}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for code, sub in df.groupby(locale, dropna=False):
        ws = wb.create_sheet(title=sanitize_sheet(code if pd.notna(code) else "unknown"))
        ws.cell(row=1, column=1, value="Link")
        ws.cell(row=1, column=2, value="nickname")
        ws.cell(row=1, column=3, value="phone")
        ws.cell(row=1, column=8, value="Requested_dep")
        ws.cell(row=1, column=9, value="Coin_Reward_Value")

        r = 2
        for _, row in sub.iterrows():
            nick  = str(row.get(nickname) or "").strip()
            loc   = str(row.get(locale)   or "ka").strip()
            amt   = normalize_amount_to_lari(row.get(reqdep))
            perc  = normalize_percent(row.get(pct))

            link = ""
            if len(amt) >= 2 and len(perc) >= 2 and amt.endswith("₾") and perc.endswith("%"):
                link = build_offer_url(loc, amt, perc, KEY)

            ws.cell(row=r, column=1, value=link)
            ws.cell(row=r, column=2, value=nick)
            ws.cell(row=r, column=3, value=row.get(phone))
            ws.cell(row=r, column=8, value=row.get(reqdep))
            ws.cell(row=r, column=9, value=row.get(coin))
            r += 1

    wb.save(out_path)
    write_log(f"SMS workbook created (URLs in col A): {out_path}")

def export_requested_dep(df, userid, reqdep, run_dir):
    _, tag = now_strings()
    out_path = os.path.join(run_dir, f"RequestedDep-{tag}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for dep_value, sub in df.groupby(reqdep, dropna=False):
        ws = wb.create_sheet(title=sanitize_sheet(dep_value if pd.notna(dep_value) else "Unknown"))
        ws.cell(row=1, column=1, value="userid")
        ws.cell(row=1, column=2, value="Requested_dep")
        r = 2
        for _, row in sub.iterrows():
            uid = str(row.get(userid) or "").strip()
            if uid:
                ws.cell(row=r, column=1, value=f"{uid},")
                ws.cell(row=r, column=2, value=row.get(reqdep))
                r += 1

    wb.save(out_path)
    write_log(f"RequestedDep workbook created: {out_path}")

# ==============================
def main():
    folder, _ = now_strings()
    run_dir = os.path.join(PROMODATE_DIR, folder)
    ensure_dirs(run_dir)
    write_log("Script started.")

    try:
        # ← Remove hardcoded usecols
        df = pd.read_excel(PROMO_FILE, dtype=object)
        cols = list(df.columns)

        userid   = resolve_column(cols, ["userid"])
        nickname = resolve_column(cols, ["nickname"])
        phone    = resolve_column(cols, ["phone"])
        pct      = resolve_column(cols, ["percentages","percentage"])
        locale   = resolve_column(cols, ["localecode","locale"])
        reqdep   = resolve_column(cols, ["requested_dep","requesteddeposit","requested"])
        coin     = resolve_column(cols, ["coin_reward_value","coin_reward"])

        if "userid" not in df.columns and userid:
            df = df.rename(columns={userid: "userid"})

        export_percentages(df, pct, run_dir)
        export_sms(df, nickname, phone, reqdep, coin, locale, pct, run_dir)
        export_requested_dep(df, "userid", reqdep, run_dir)

        write_log("Script finished successfully.")
        print("Done.")
    except Exception as e:
        write_log(f"Error: {e}")
        raise

if __name__ == "__main__":
    main()
